"""
MCA Bank Statement Extractor - MoneyThumb Compatible Output
Extracts and structures data exactly like MoneyThumb for underwriting

Output Structure (mirrors MoneyThumb xlsx):
1. Revenue Statistics - Monthly/Annual metrics
2. Statements Summary - Per-account breakdown  
3. Monthly MCA - MCA by month
4. MCA Companies - Grouped by lender
5. MCA Transactions - Individual MCA hits
6. NSF Transactions
7. Overdraft Transactions
8. Daily Balances (with True Balance)
9. Daily Cash Flows
10. Monthly Cash Flows
11. Credit Transactions (all deposits)
12. True Credit Transactions (real revenue only)
13. Non-True Credit Transactions (loans, transfers)
14. Incoming Transfers
15. Outgoing Transfers
16. Large Transactions (>$1000)
17. Returned Transactions
18. Monthly Negative Days
19. Repeating Transactions
"""

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

from pathlib import Path
import json
import re
import csv
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
from dataclasses import dataclass, field, asdict
from collections import defaultdict
import calendar


# Known MCA Lenders - Pattern matching for ACH descriptions
MCA_LENDERS = {
    'Kapitus': [r'KAPITUS', r'KAPTIUS'],
    'Forward Financing': [r'FORWARD\s*FINANCING', r'FORWARD\s*FIN'],
    'Family Funding': [r'FAMILY\s*FUNDING', r'FAMILYFUND'],
    'Credibly': [r'CREDIBLY'],
    'Clearco': [r'CLEARCO'],
    'Fundbox': [r'FUNDBOX'],
    'Bluevine': [r'BLUEVINE', r'BLUE\s*VINE'],
    'OnDeck': [r'ONDECK', r'ON\s*DECK'],
    'Rapid Finance': [r'RAPID\s*FINANCE', r'RAPIDFIN'],
    'National Funding': [r'NATIONAL\s*FUNDING'],
    'Intuit Financing': [r'INTUIT\s*FINANCING', r'QBC_PMTS\s*INTUIT', r'WEBBANK/INTUIT'],
    'Funders App': [r'FUNDERS\s*APP'],
    'Shopify Capital': [r'SHOPIFY\s*CAPITAL'],
    'PayPal Working Capital': [r'PAYPAL\s*WORKING', r'PPWC'],
    'Square Capital': [r'SQUARE\s*CAPITAL', r'SQ\s*CAPITAL'],
    'Amazon Lending': [r'AMAZON\s*LENDING'],
    'Libertas Funding': [r'LIBERTAS'],
    'Merchant Cash': [r'MERCHANT\s*CASH'],
    'Business Backer': [r'BUSINESS\s*BACKER'],
    'Yellowstone Capital': [r'YELLOWSTONE'],
    'World Business Lenders': [r'WORLD\s*BUSINESS', r'WBL'],
    'Lendr': [r'LENDR\b'],
    'Lendio': [r'LENDIO'],
    'Kabbage': [r'KABBAGE'],
    'Can Capital': [r'CAN\s*CAPITAL'],
    'Behalf': [r'\bBEHALF\b'],
    'Breakout Capital': [r'BREAKOUT'],
    'Reliant Funding': [r'RELIANT'],
    'Mulligan Funding': [r'MULLIGAN'],
    'CFG Merchant': [r'CFG\s*MERCHANT'],
    'Everest Business': [r'EVEREST\s*BUSINESS'],
    'Billd Exchange': [r'BILLD\s*EXCHANGE', r'BILLD'],
    'Cashera Private': [r'CASHERA\s*PRIVATE', r'CASHERA'],
}

# Patterns for non-true revenue (loans, transfers, not real income)
NON_TRUE_PATTERNS = [
    r'LOAN',
    r'XFER',
    r'TRANSFER',
    r'WEBBANK',
    r'LENDING',
    r'CAPITAL\s*ONE.*DEPOSIT',  # Capital injection not revenue
]

# Patterns for incoming transfers
TRANSFER_IN_PATTERNS = [
    r'TRANSFER\s*CREDIT',
    r'XFER\s*CR',
    r'OL\s*XFER',
    r'WIRE\s*CR',
    r'INCOMING\s*WIRE',
]

# Patterns for outgoing transfers
TRANSFER_OUT_PATTERNS = [
    r'TRANSFER\s*DEBIT',
    r'XFER\s*DB',
    r'OL\s*XFER',
    r'WIRE\s*DB',
    r'OUTGOING\s*WIRE',
    r'INST\s*XFER',
]


@dataclass
class Transaction:
    """Individual transaction"""
    account: str
    date: datetime
    description: str
    amount: float
    memo: str = ""
    number: str = "0"
    type: str = ""
    
    def is_credit(self) -> bool:
        return self.amount > 0
    
    def is_debit(self) -> bool:
        return self.amount < 0


@dataclass 
class DailyBalance:
    """Daily balance record"""
    date: datetime
    balance: float
    true_balance: float  # Excluding loan deposits


@dataclass
class MCATransaction:
    """MCA-specific transaction"""
    lender: str
    account: str
    date: datetime
    description: str
    amount: float
    memo: str = ""
    number: str = "0"
    type: str = ""


@dataclass
class MonthlyMCA:
    """MCA activity for a month"""
    month: str
    work_days: int
    account: str
    lender: str
    withdrawal_count: int
    withdrawal_total: float
    deposit_total: float
    deposit_dates: str
    latest_withdrawal_amount: float


@dataclass
class MCACompany:
    """MCA company aggregate"""
    lender: str
    month: str
    work_days: int
    term: str
    withdrawal_count: int
    withdrawal_total: float
    withhold_percent: float
    deposit_count: int
    deposit_total: float
    withdrawal_frequency: str
    last_withdrawal_date: datetime
    deposit_dates: str
    last_withdrawal_amount: float


@dataclass
class StatementSummary:
    """Per-statement summary (one row per month)"""
    account: str
    bank_name: str
    statement_month: str
    starting_balance: float
    total_credits: float
    num_credits: int
    true_credits: float
    num_true_credits: int
    total_debits: float
    num_debits: int
    ending_balance: float
    avg_balance: float
    avg_true_balance: float
    days_negative: int
    num_overdrafts: int
    num_nsfs: int
    low_days: int
    mca_withhold_percent: float


@dataclass
class RevenueStatistics:
    """Top-level revenue stats (MoneyThumb Revenue Statistics sheet)"""
    # Monthly figures
    revenue_monthly: float = 0.0
    true_revenue_monthly: float = 0.0
    non_true_revenue_monthly: float = 0.0
    expenses_monthly: float = 0.0
    profit_monthly: float = 0.0
    
    # Annual figures (monthly * 12)
    revenue_annual: float = 0.0
    true_revenue_annual: float = 0.0
    non_true_revenue_annual: float = 0.0
    expenses_annual: float = 0.0
    profit_annual: float = 0.0
    
    # Balance metrics
    combined_avg_daily_balance: float = 0.0
    combined_avg_daily_true_balance: float = 0.0
    days_negative: int = 0
    combined_days_negative: int = 0
    low_days: int = 0
    
    # MCA metrics
    mca_withhold_percent: float = 0.0
    total_debt_withdrawals: float = 0.0
    dti_percent: float = 0.0
    
    # Additional
    min_monthly_true_revenue: float = 0.0
    avg_monthly_net_revenue: float = 0.0
    avg_monthly_factoring_revenue: float = 0.0
    avg_monthly_credit_card_revenue: float = 0.0


@dataclass
class BankStatementAnalysis:
    """Complete analysis matching MoneyThumb output"""
    # Metadata
    business_name: str = ""
    address: str = ""
    
    # Core data structures
    revenue_statistics: RevenueStatistics = field(default_factory=RevenueStatistics)
    statement_summaries: List[StatementSummary] = field(default_factory=list)
    
    # All transactions
    all_transactions: List[Transaction] = field(default_factory=list)
    credit_transactions: List[Transaction] = field(default_factory=list)
    true_credit_transactions: List[Transaction] = field(default_factory=list)
    non_true_credit_transactions: List[Transaction] = field(default_factory=list)
    
    # Transfers
    incoming_transfers: List[Transaction] = field(default_factory=list)
    outgoing_transfers: List[Transaction] = field(default_factory=list)
    
    # MCA data
    mca_transactions: List[MCATransaction] = field(default_factory=list)
    monthly_mca: List[MonthlyMCA] = field(default_factory=list)
    mca_companies: List[MCACompany] = field(default_factory=list)
    
    # Risk indicators
    nsf_transactions: List[Transaction] = field(default_factory=list)
    overdraft_transactions: List[Transaction] = field(default_factory=list)
    returned_transactions: List[Transaction] = field(default_factory=list)
    
    # Balances and flows
    daily_balances: List[DailyBalance] = field(default_factory=list)
    daily_cash_flows: List[Dict] = field(default_factory=list)
    monthly_cash_flows: List[Dict] = field(default_factory=list)
    monthly_negative_days: List[Dict] = field(default_factory=list)
    
    # Other
    large_transactions: List[Transaction] = field(default_factory=list)
    repeating_transactions: List[Transaction] = field(default_factory=list)


def is_native_pdf(pdf_path: str) -> bool:
    """Detect if PDF has extractable text or needs OCR"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text_chars = 0
            pages_checked = min(3, len(pdf.pages))
            for i in range(pages_checked):
                text = pdf.pages[i].extract_text() or ""
                text_chars += len(text.strip())
            return (text_chars / pages_checked) > 100 if pages_checked > 0 else False
    except Exception:
        return False


def extract_text_native(pdf_path: str) -> str:
    """Extract text from native PDF using pdfplumber"""
    full_text = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text.append(text)
    return "\n\n".join(full_text)


def extract_text_ocr(pdf_path: str) -> str:
    """Extract text using Surya OCR for scanned PDFs"""
    try:
        from surya.ocr import run_ocr
        from surya.model.detection.model import load_model as load_det_model
        from surya.model.recognition.model import load_model as load_rec_model
        from surya.model.detection.processor import load_processor as load_det_processor
        from surya.model.recognition.processor import load_processor as load_rec_processor
        from PIL import Image
        
        print("Loading Surya OCR models...")
        det_model = load_det_model()
        det_processor = load_det_processor()
        rec_model = load_rec_model()
        rec_processor = load_rec_processor()
        
        doc = fitz.open(pdf_path)
        images = []
        for page_num in range(len(doc)):
            page = doc[page_num]
            mat = fitz.Matrix(300/72, 300/72)
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            images.append(img)
        doc.close()
        
        print(f"Running OCR on {len(images)} pages...")
        results = run_ocr(images, [["en"]] * len(images), 
                         det_model, det_processor, rec_model, rec_processor)
        
        full_text = []
        for page_result in results:
            page_text = [line.text for line in page_result.text_lines]
            full_text.append("\n".join(page_text))
        return "\n\n".join(full_text)
    except ImportError:
        print("Surya OCR not available")
        return ""


def parse_currency(text: str) -> float:
    """Parse currency string to float"""
    if not text:
        return 0.0
    cleaned = re.sub(r'[,$\s]', '', str(text))
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = '-' + cleaned[1:-1]
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_date(text: str) -> Optional[datetime]:
    """Parse various date formats"""
    formats = [
        '%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%m-%d-%Y',
        '%B %d, %Y', '%b %d, %Y', '%d-%b-%Y'
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text.strip(), fmt)
        except ValueError:
            continue
    return None


def detect_mca_lender(description: str) -> Optional[str]:
    """Detect MCA lender from transaction description"""
    desc_upper = description.upper()
    for lender, patterns in MCA_LENDERS.items():
        for pattern in patterns:
            if re.search(pattern, desc_upper):
                return lender
    return None


def is_non_true_credit(description: str) -> bool:
    """Check if credit is non-true revenue (loan, transfer, etc)"""
    desc_upper = description.upper()
    for pattern in NON_TRUE_PATTERNS:
        if re.search(pattern, desc_upper):
            return True
    return False


def is_incoming_transfer(description: str) -> bool:
    """Check if transaction is an incoming transfer"""
    desc_upper = description.upper()
    for pattern in TRANSFER_IN_PATTERNS:
        if re.search(pattern, desc_upper):
            return True
    return False


def is_outgoing_transfer(description: str) -> bool:
    """Check if transaction is an outgoing transfer"""
    desc_upper = description.upper()
    for pattern in TRANSFER_OUT_PATTERNS:
        if re.search(pattern, desc_upper):
            return True
    return False


def get_work_days_in_month(year: int, month: int) -> int:
    """Calculate business days in a month"""
    cal = calendar.Calendar()
    work_days = 0
    for day in cal.itermonthdays2(year, month):
        if day[0] != 0 and day[1] < 5:  # Not weekend
            work_days += 1
    return work_days


def parse_transactions_csv(csv_path: str) -> List[Transaction]:
    """Parse MoneyThumb CSV transaction export"""
    transactions = []
    current_account = ""
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader, None)
        
        for row in reader:
            if len(row) < 5:
                continue
            
            # Check for account header row
            if row[0] and row[0].startswith('Account'):
                match = re.search(r'(xxxx\d+)', row[0])
                if match:
                    current_account = match.group(1)
                continue
            
            # Parse transaction row
            try:
                date_str = row[1] if len(row) > 1 else ""
                description = row[2] if len(row) > 2 else ""
                amount_str = row[3] if len(row) > 3 else "0"
                memo = row[4] if len(row) > 4 else ""
                number = row[5] if len(row) > 5 else "0"
                txn_type = row[6] if len(row) > 6 else ""
                
                date = parse_date(date_str)
                if not date or not description:
                    continue
                    
                amount = parse_currency(amount_str)
                
                transactions.append(Transaction(
                    account=current_account,
                    date=date,
                    description=description,
                    amount=amount,
                    memo=memo,
                    number=number,
                    type=txn_type
                ))
            except Exception as e:
                continue
    
    return transactions


def analyze_transactions(transactions: List[Transaction]) -> BankStatementAnalysis:
    """Analyze transactions and build complete MoneyThumb-compatible output"""
    analysis = BankStatementAnalysis()
    analysis.all_transactions = transactions
    
    if not transactions:
        return analysis
    
    # Sort by date
    transactions.sort(key=lambda t: t.date)
    
    # Separate credits and debits
    credits = [t for t in transactions if t.is_credit()]
    debits = [t for t in transactions if t.is_debit()]
    
    # Classify credits
    for t in credits:
        analysis.credit_transactions.append(t)
        
        if is_non_true_credit(t.description):
            analysis.non_true_credit_transactions.append(t)
        else:
            analysis.true_credit_transactions.append(t)
        
        if is_incoming_transfer(t.description):
            analysis.incoming_transfers.append(t)
    
    # Classify debits and detect MCAs
    for t in debits:
        lender = detect_mca_lender(t.description)
        if lender:
            mca_txn = MCATransaction(
                lender=lender,
                account=t.account,
                date=t.date,
                description=t.description,
                amount=t.amount,
                memo=t.memo,
                number=t.number,
                type=t.type
            )
            analysis.mca_transactions.append(mca_txn)
        
        if is_outgoing_transfer(t.description):
            analysis.outgoing_transfers.append(t)
        
        # Check for NSF/Overdraft
        desc_upper = t.description.upper()
        if 'NSF' in desc_upper or 'INSUFFICIENT' in desc_upper:
            analysis.nsf_transactions.append(t)
        if 'OVERDRAFT' in desc_upper or 'OD FEE' in desc_upper:
            analysis.overdraft_transactions.append(t)
        if 'RETURN' in desc_upper and 'CHECK' in desc_upper:
            analysis.returned_transactions.append(t)
    
    # Large transactions (>$1000)
    analysis.large_transactions = [t for t in transactions if abs(t.amount) >= 1000]
    
    # Group MCA by month and lender
    mca_by_month_lender = defaultdict(lambda: defaultdict(list))
    for mca in analysis.mca_transactions:
        month_key = mca.date.strftime('%B %Y')
        mca_by_month_lender[month_key][mca.lender].append(mca)
    
    # Build Monthly MCA and MCA Companies
    for month_key, lenders in mca_by_month_lender.items():
        for lender, mca_list in lenders.items():
            withdrawals = [m for m in mca_list if m.amount < 0]
            deposits = [m for m in mca_list if m.amount > 0]
            
            # Get work days
            sample_date = mca_list[0].date
            work_days = get_work_days_in_month(sample_date.year, sample_date.month)
            
            withdrawal_total = sum(m.amount for m in withdrawals)
            deposit_total = sum(m.amount for m in deposits)
            deposit_dates = ', '.join(sorted(set(m.date.strftime('%Y-%m-%d') for m in deposits)))
            
            latest_withdrawal = max(withdrawals, key=lambda m: m.date) if withdrawals else None
            
            monthly = MonthlyMCA(
                month=month_key,
                work_days=work_days,
                account=mca_list[0].account,
                lender=lender,
                withdrawal_count=len(withdrawals),
                withdrawal_total=withdrawal_total,
                deposit_total=deposit_total,
                deposit_dates=deposit_dates,
                latest_withdrawal_amount=latest_withdrawal.amount if latest_withdrawal else 0
            )
            analysis.monthly_mca.append(monthly)
    
    # Calculate revenue statistics
    total_credits = sum(t.amount for t in analysis.credit_transactions)
    true_credits = sum(t.amount for t in analysis.true_credit_transactions)
    non_true_credits = sum(t.amount for t in analysis.non_true_credit_transactions)
    total_debits = sum(abs(t.amount) for t in debits)
    mca_total = sum(abs(m.amount) for m in analysis.mca_transactions if m.amount < 0)
    
    # Get unique months
    months = set(t.date.strftime('%Y-%m') for t in transactions)
    num_months = len(months) or 1
    
    stats = analysis.revenue_statistics
    stats.revenue_monthly = total_credits / num_months
    stats.true_revenue_monthly = true_credits / num_months
    stats.non_true_revenue_monthly = non_true_credits / num_months
    stats.expenses_monthly = total_debits / num_months
    stats.profit_monthly = stats.true_revenue_monthly - stats.expenses_monthly
    
    stats.revenue_annual = stats.revenue_monthly * 12
    stats.true_revenue_annual = stats.true_revenue_monthly * 12
    stats.non_true_revenue_annual = stats.non_true_revenue_monthly * 12
    stats.expenses_annual = stats.expenses_monthly * 12
    stats.profit_annual = stats.profit_monthly * 12
    
    stats.total_debt_withdrawals = mca_total
    if stats.true_revenue_monthly > 0:
        stats.mca_withhold_percent = (mca_total / num_months) / stats.true_revenue_monthly
    
    # Find repeating transactions (same description, multiple occurrences)
    desc_counts = defaultdict(list)
    for t in transactions:
        # Normalize description for matching
        key = re.sub(r'\d{6,}', 'XXXX', t.description)  # Remove long numbers
        desc_counts[key].append(t)
    
    for key, txn_list in desc_counts.items():
        if len(txn_list) >= 3:  # At least 3 occurrences = repeating
            analysis.repeating_transactions.extend(txn_list)
    
    return analysis


def export_to_xlsx(analysis: BankStatementAnalysis, output_path: str):
    """Export analysis to MoneyThumb-compatible xlsx"""
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return
    
    wb = Workbook()
    
    # Sheet 1: Revenue Statistics
    ws = wb.active
    ws.title = "Revenue Statistics"
    ws.append(['Label', 'Monthly', 'Annual'])
    stats = analysis.revenue_statistics
    ws.append(['Revenue', stats.revenue_monthly, stats.revenue_annual])
    ws.append(['True Revenue', stats.true_revenue_monthly, stats.true_revenue_annual])
    ws.append(['Expenses', stats.expenses_monthly, stats.expenses_annual])
    ws.append(['Profit', stats.profit_monthly, stats.profit_annual])
    ws.append(['Balance/Days Negative', stats.combined_avg_daily_balance, stats.days_negative])
    ws.append(['Non-True Revenue', stats.non_true_revenue_monthly, stats.non_true_revenue_annual])
    ws.append(['MCA Withhold Percent', stats.mca_withhold_percent, 0])
    ws.append(['Total Debt Withdrawals', stats.total_debt_withdrawals, 0])
    
    # Sheet 2: MCA Transactions
    ws = wb.create_sheet("MCA Transactions")
    ws.append(['Lender', 'Account', 'Date', 'Description', 'Amount', 'Memo', 'Number', 'Type'])
    for m in analysis.mca_transactions:
        ws.append([m.lender, m.account, m.date, m.description, m.amount, m.memo, m.number, m.type])
    
    # Sheet 3: Monthly MCA
    ws = wb.create_sheet("Monthly MCA")
    ws.append(['Month', 'Work Days', 'Account', 'Lender', 'Withdrawal Count', 
               'Withdrawal Total', 'Deposit Total', 'Deposit Dates', 'Latest Withdrawal Amount'])
    for m in analysis.monthly_mca:
        ws.append([m.month, m.work_days, m.account, m.lender, m.withdrawal_count,
                  m.withdrawal_total, m.deposit_total, m.deposit_dates, m.latest_withdrawal_amount])
    
    # Sheet 4: Credit Transactions
    ws = wb.create_sheet("Credit Transactions")
    ws.append(['Account', 'Date', 'Description', 'Amount', 'Memo', 'Number', 'Type'])
    for t in analysis.credit_transactions:
        ws.append([t.account, t.date, t.description, t.amount, t.memo, t.number, t.type])
    
    # Sheet 5: True Credit Transactions
    ws = wb.create_sheet("True Credit Transactions")
    ws.append(['Account', 'Date', 'Description', 'Amount', 'Memo', 'Number', 'Type'])
    for t in analysis.true_credit_transactions:
        ws.append([t.account, t.date, t.description, t.amount, t.memo, t.number, t.type])
    
    # Sheet 6: Non-True Credit Transactions
    ws = wb.create_sheet("Non-True Credit Transactions")
    ws.append(['Account', 'Date', 'Description', 'Amount', 'Memo', 'Number', 'Type'])
    for t in analysis.non_true_credit_transactions:
        ws.append([t.account, t.date, t.description, t.amount, t.memo, t.number, t.type])
    
    # Sheet 7: NSF Transactions
    ws = wb.create_sheet("NSF Transactions")
    ws.append(['Account', 'Date', 'Description', 'Amount', 'Memo', 'Number', 'Type'])
    for t in analysis.nsf_transactions:
        ws.append([t.account, t.date, t.description, t.amount, t.memo, t.number, t.type])
    
    # Sheet 8: Overdraft Transactions
    ws = wb.create_sheet("Overdraft Transactions")
    ws.append(['Account', 'Date', 'Description', 'Amount', 'Memo', 'Number', 'Type'])
    for t in analysis.overdraft_transactions:
        ws.append([t.account, t.date, t.description, t.amount, t.memo, t.number, t.type])
    
    # Sheet 9: Incoming Transfers
    ws = wb.create_sheet("Incoming Transfers")
    ws.append(['Account', 'Date', 'Description', 'Amount', 'Memo', 'Number', 'Type'])
    for t in analysis.incoming_transfers:
        ws.append([t.account, t.date, t.description, t.amount, t.memo, t.number, t.type])
    
    # Sheet 10: Outgoing Transfers
    ws = wb.create_sheet("Outgoing Transfers")
    ws.append(['Account', 'Date', 'Description', 'Amount', 'Memo', 'Number', 'Type'])
    for t in analysis.outgoing_transfers:
        ws.append([t.account, t.date, t.description, t.amount, t.memo, t.number, t.type])
    
    # Sheet 11: Large Transactions
    ws = wb.create_sheet("Large Transactions")
    ws.append(['Account', 'Date', 'Description', 'Amount', 'Memo', 'Number', 'Type'])
    for t in analysis.large_transactions:
        ws.append([t.account, t.date, t.description, t.amount, t.memo, t.number, t.type])
    
    # Sheet 12: Returned Transactions
    ws = wb.create_sheet("Returned Transactions")
    ws.append(['Account', 'Date', 'Description', 'Amount', 'Memo', 'Number', 'Type'])
    for t in analysis.returned_transactions:
        ws.append([t.account, t.date, t.description, t.amount, t.memo, t.number, t.type])
    
    # Sheet 13: Repeating Transactions
    ws = wb.create_sheet("Repeating Transactions")
    ws.append(['Account', 'Date', 'Description', 'Amount', 'Memo', 'Number', 'Type'])
    for t in analysis.repeating_transactions:
        ws.append([t.account, t.date, t.description, t.amount, t.memo, t.number, t.type])
    
    wb.save(output_path)
    print(f"Saved: {output_path}")


def export_to_json(analysis: BankStatementAnalysis, output_path: str):
    """Export analysis to JSON for API/database use"""
    
    def serialize(obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif hasattr(obj, '__dict__'):
            return {k: serialize(v) for k, v in obj.__dict__.items()}
        elif isinstance(obj, list):
            return [serialize(item) for item in obj]
        elif isinstance(obj, dict):
            return {k: serialize(v) for k, v in obj.items()}
        return obj
    
    data = serialize(analysis)
    
    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2, default=str)
    
    print(f"Saved: {output_path}")


def print_summary(analysis: BankStatementAnalysis):
    """Print analysis summary"""
    stats = analysis.revenue_statistics
    
    print("\n" + "=" * 60)
    print("BANK STATEMENT ANALYSIS - MONEYTHUMB FORMAT")
    print("=" * 60)
    
    print(f"\n📊 REVENUE STATISTICS (Monthly)")
    print(f"   Total Revenue:      ${stats.revenue_monthly:>12,.2f}")
    print(f"   True Revenue:       ${stats.true_revenue_monthly:>12,.2f}")
    print(f"   Non-True Revenue:   ${stats.non_true_revenue_monthly:>12,.2f}")
    print(f"   Total Expenses:     ${stats.expenses_monthly:>12,.2f}")
    print(f"   Profit:             ${stats.profit_monthly:>12,.2f}")
    
    print(f"\n📊 REVENUE STATISTICS (Annual)")
    print(f"   Total Revenue:      ${stats.revenue_annual:>12,.2f}")
    print(f"   True Revenue:       ${stats.true_revenue_annual:>12,.2f}")
    
    if analysis.mca_transactions:
        print(f"\n⚠️  MCA POSITIONS DETECTED: {len(set(m.lender for m in analysis.mca_transactions))} lenders")
        
        # Group by lender
        by_lender = defaultdict(list)
        for m in analysis.mca_transactions:
            by_lender[m.lender].append(m)
        
        for lender, txns in by_lender.items():
            total = sum(t.amount for t in txns)
            withdrawals = [t for t in txns if t.amount < 0]
            deposits = [t for t in txns if t.amount > 0]
            print(f"   {lender}:")
            print(f"      Withdrawals: {len(withdrawals)} totaling ${abs(sum(t.amount for t in withdrawals)):,.2f}")
            if deposits:
                print(f"      Deposits: {len(deposits)} totaling ${sum(t.amount for t in deposits):,.2f}")
        
        print(f"\n   MCA Withhold %: {stats.mca_withhold_percent:.1%}")
    
    print(f"\n🚨 RISK INDICATORS")
    print(f"   NSF Transactions: {len(analysis.nsf_transactions)}")
    print(f"   Overdraft Transactions: {len(analysis.overdraft_transactions)}")
    print(f"   Returned Transactions: {len(analysis.returned_transactions)}")
    
    print(f"\n📈 TRANSACTION COUNTS")
    print(f"   Total Transactions: {len(analysis.all_transactions)}")
    print(f"   Credit Transactions: {len(analysis.credit_transactions)}")
    print(f"   True Credits: {len(analysis.true_credit_transactions)}")
    print(f"   Large Transactions (>$1K): {len(analysis.large_transactions)}")
    print(f"   Repeating Transactions: {len(analysis.repeating_transactions)}")


def process_csv(csv_path: str, output_dir: str = "output") -> BankStatementAnalysis:
    """Process MoneyThumb CSV export and generate analysis"""
    print(f"\nProcessing: {csv_path}")
    
    transactions = parse_transactions_csv(csv_path)
    print(f"Parsed {len(transactions)} transactions")
    
    analysis = analyze_transactions(transactions)
    print_summary(analysis)
    
    # Export
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    csv_name = Path(csv_path).stem
    export_to_xlsx(analysis, str(output_path / f"{csv_name}_analysis.xlsx"))
    export_to_json(analysis, str(output_path / f"{csv_name}_analysis.json"))
    
    return analysis


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python moneythumb_extractor.py <transactions.csv> [output_dir]")
        print("\nExample:")
        print("  python moneythumb_extractor.py Application_FA42148_transactions.csv output/")
        sys.exit(1)
    
    csv_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "output"
    
    process_csv(csv_path, output_dir)
